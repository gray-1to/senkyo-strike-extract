from datetime import date
from typing import Union

import pandas as pd
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm
from utils.os.output import prepare_output_pd


def operation(input_ws: Union[Worksheet, ReadOnlyWorksheet], progressReport: bool = False) -> list[pd.DataFrame]:

    # input_ws.fillna("", inplace=True)

    input_width: int = input_ws.max_column
    input_height: int = input_ws.max_row

    # 出力データ用配列を準備
    strike_data: pd.DataFrame = prepare_output_pd(input_height, input_width, input_ws)
    no_strike_data: pd.DataFrame = prepare_output_pd(input_height, input_width, input_ws)

    # check対象の列を特定
    check_col_indexes: list[int] = []
    for col_index in range(0, input_width):
        # openpyxlのcellは1列目はcolumnが1になるので+1
        if input_ws.cell(row=1, column=col_index + 1).value == "check":
            check_col_indexes.insert(-1, col_index)

    # 探索準備
    strike_data_height: int = 2
    no_strike_data_height: int = 2
    # 探索開始
    # 1,2行目はヘッダーなので、3行目からチェック開始
    if progressReport:
        # for --progress-report option
        for row_index in tqdm(range(2, input_height)):
            strike_detected: bool = False
            # check対象の列をチェック
            for col_index in check_col_indexes:
                # openpyxlのcellは1列目はcolumnが1になるので+1
                if input_ws.cell(row=row_index + 1, column=col_index + 1).font.strike is True:
                    strike_detected = True
            # 打ち消し線がcheck対象の列にある場合
            if strike_detected:
                # strike.xlsxの最終行に追加
                for copy_to_col_index in range(0, input_width):
                    original_value: Union[date, str, int, None] = input_ws.cell(
                        row=row_index + 1, column=copy_to_col_index + 1
                    ).value  # openpyxlのcellは1列目はcolumnが1になるので+1
                    strike_data[copy_to_col_index][strike_data_height] = original_value
                strike_data_height = strike_data_height + 1
            # 打ち消し線がcheck対象の列にない場合
            else:
                # no_strike.xlsxの最終行に追加
                for copy_to_col_index in range(0, input_width):
                    original_value: Union[date, str, int, None] = input_ws.cell(
                        row=row_index + 1, column=copy_to_col_index + 1
                    ).value  # openpyxlのcellは1列目はcolumnが1になるので+1
                    no_strike_data[copy_to_col_index][no_strike_data_height] = original_value
                no_strike_data_height = no_strike_data_height + 1
    else:
        # for no --progress-report option
        for row_index in range(2, input_height):
            strike_detected: bool = False
            # check対象の列をチェック
            for col_index in check_col_indexes:
                # openpyxlのcellは1列目はcolumnが1になるので+1
                if input_ws.cell(row=row_index + 1, column=col_index + 1).font.strike is True:
                    strike_detected = True
            # 打ち消し線がcheck対象の列にある場合
            if strike_detected:
                # strike.xlsxの最終行に追加
                for copy_to_col_index in range(0, input_width):
                    original_value: Union[date, str, int, None] = input_ws.cell(
                        row=row_index + 1, column=copy_to_col_index + 1
                    ).value  # openpyxlのcellは1列目はcolumnが1になるので+1
                    strike_data[copy_to_col_index][strike_data_height] = original_value
                strike_data_height = strike_data_height + 1
            # 打ち消し線がcheck対象の列にない場合
            else:
                # no_strike.xlsxの最終行に追加
                for copy_to_col_index in range(0, input_width):
                    original_value: Union[date, str, int, None] = input_ws.cell(
                        row=row_index + 1, column=copy_to_col_index + 1
                    ).value  # openpyxlのcellは1列目はcolumnが1になるので+1
                    no_strike_data[copy_to_col_index][no_strike_data_height] = original_value
                no_strike_data_height = no_strike_data_height + 1

    return [strike_data, no_strike_data]
